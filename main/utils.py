import csv
import os
from io import StringIO
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from .models import Customer


def normalize_column_name(column_name):
    """Normalize column names to handle variations from 1C exports"""
    if not column_name:
        return None
    
    column_lower = column_name.lower().strip()
    
    # Map common variations to standard field names
    name_mappings = {
        'name': ['имя', 'name', 'firstname', 'first name', 'имя клиента', 'название'],
        'surname': ['фамилия', 'surname', 'lastname', 'last name', 'фамилия клиента', 'last'],
        'place': ['место', 'place', 'город', 'city', 'location', 'адрес', 'address', 'откуда'],
        'phone': ['телефон', 'phone', 'tel', 'телефон клиента', 'phone number'],
        'address': ['адрес', 'address', 'полный адрес', 'full address', 'адрес клиента'],
        'counterparty': ['контрагент', 'counterparty', 'клиент', 'client', 'customer', 'покупатель'],
    }
    
    for field, variations in name_mappings.items():
        if column_lower in variations or any(v in column_lower for v in variations):
            return field
    
    return None


def parse_counterparty_name(counterparty_value):
    """
    Parse Контрагент (Counterparty) field to extract name and surname.
    Handles formats like:
    - "Иванов Иван" (surname first)
    - "Иван Иванов" (name first)
    - "Иванов Иван Петрович" (surname, name, patronymic) - MOST COMMON IN 1C
    - Just one name
    - Empty or None values
    - Numbers or special characters
    
    Returns: (name, surname, patronymic) or (name, surname, None)
    """
    if not counterparty_value:
        return None, None, None
    
    # Clean the value - convert to string and strip
    full_name = str(counterparty_value).strip()
    
    # Remove extra whitespace
    full_name = ' '.join(full_name.split())
    
    if not full_name or full_name.lower() in ['none', 'null', 'nan', '']:
        return None, None, None
    
    # Check if it's just a number or special characters (not a name)
    if full_name.replace('.', '').replace(',', '').replace('-', '').replace(' ', '').isdigit():
        return None, None, None
    
    # Split by spaces
    parts = [p.strip() for p in full_name.split() if p.strip()]
    
    if len(parts) == 0:
        return None, None, None
    elif len(parts) == 1:
        # Only one part - use it as surname
        return None, parts[0], None
    elif len(parts) == 2:
        # Two parts - typically "Surname Name" in Russian
        # We'll assume first is surname, second is name
        return parts[1], parts[0], None  # name, surname, patronymic
    else:
        # Three or more parts - assume: Surname Name Patronymic ...
        # This is the most common format in 1C: "Иванов Иван Петрович"
        surname = parts[0]
        name = parts[1]
        patronymic = ' '.join(parts[2:]) if len(parts) > 2 else None  # Join remaining parts as patronymic
        return name, surname, patronymic


def parse_csv_file(file):
    """Parse CSV file and return list of dictionaries"""
    try:
        # Reset file pointer to beginning
        file.seek(0)
        # Try to detect encoding
        content = file.read()
        
        # Try different encodings (1C often uses Windows-1251 or UTF-8)
        encodings = ['utf-8-sig', 'utf-8', 'windows-1251', 'cp1251', 'latin-1']
        decoded_content = None
        
        for encoding in encodings:
            try:
                decoded_content = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if decoded_content is None:
            raise ValueError("Could not decode file. Please ensure it's a valid CSV file.")
        
        # Parse CSV
        csv_reader = csv.DictReader(StringIO(decoded_content))
        rows = list(csv_reader)
        
        if not rows:
            raise ValueError("CSV file is empty or has no data rows.")
        
        # Normalize column names
        normalized_rows = []
        for row in rows:
            normalized_row = {}
            for key, value in row.items():
                normalized_key = normalize_column_name(key)
                if normalized_key:
                    normalized_row[normalized_key] = value.strip() if value else ''
                else:
                    # Keep original key if we can't normalize it
                    normalized_row[key] = value.strip() if value else ''
            normalized_rows.append(normalized_row)
        
        return normalized_rows
    
    except Exception as e:
        raise ValueError(f"Error parsing CSV file: {str(e)}")


def parse_excel_file(file):
    """Parse Excel file and return list of dictionaries"""
    # Try to import openpyxl with better error handling
    try:
        import openpyxl
    except ImportError as e:
        import sys
        python_path = sys.executable
        error_msg = (
            f"openpyxl is required for Excel file support.\n"
            f"Current Python: {python_path}\n"
            f"Please install it by running: pip install openpyxl\n"
            f"Make sure you're using the correct Python environment (activate your virtual environment first).\n"
            f"Original error: {str(e)}"
        )
        raise ImportError(error_msg)
    
    try:
        # Reset file pointer to beginning
        file.seek(0)
        # Try to load workbook - handle both file objects and file paths
        if hasattr(file, 'read'):
            workbook = openpyxl.load_workbook(file, data_only=True)
        else:
            workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value if cell.value else '')
        
        # Normalize headers and keep original for counterparty
        normalized_headers = []
        header_map = {}
        original_headers = []
        for i, header in enumerate(headers):
            original_header = str(header) if header else ''
            original_headers.append(original_header)
            normalized = normalize_column_name(original_header)
            normalized_headers.append(normalized)
            if normalized:
                header_map[normalized] = i
        
        # Read data rows
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue
            
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(normalized_headers):
                    normalized_key = normalized_headers[i]
                    orig_key = original_headers[i] if i < len(original_headers) else ''
                    
                    if normalized_key:
                        # Convert value to string, handling None and numbers
                        if value is None:
                            str_value = ''
                        elif isinstance(value, (int, float)):
                            str_value = str(value)
                        else:
                            str_value = str(value).strip() if value else ''
                        row_dict[normalized_key] = str_value
                    elif orig_key:
                        # Keep original column name for unrecognized columns
                        # Also check if it's Контрагент in Russian
                        if 'контрагент' in orig_key.lower() or 'counterparty' in orig_key.lower():
                            # This is the counterparty column, normalize it
                            if value is None:
                                str_value = ''
                            elif isinstance(value, (int, float)):
                                str_value = str(value)
                            else:
                                str_value = str(value).strip() if value else ''
                            row_dict['counterparty'] = str_value
                        else:
                            row_dict[orig_key] = str(value).strip() if value else ''
            
            # Handle counterparty field - parse it into name, surname, and patronymic
            if 'counterparty' in row_dict:
                counterparty_value = row_dict.get('counterparty', '')
                # Convert to string and clean - handle None, empty strings, and numbers
                if counterparty_value is not None:
                    counterparty_str = str(counterparty_value).strip()
                    # Check if it's a number (like 0.0, 100, etc.) - these are not names
                    try:
                        float(counterparty_str)
                        # It's a number, skip this row_dict
                        counterparty_str = ''
                    except ValueError:
                        # Not a number, proceed
                        pass
                    
                    if counterparty_str and counterparty_str.lower() not in ['none', 'null', 'nan', '']:
                        name, surname, patronymic = parse_counterparty_name(counterparty_str)
                        if name:
                            row_dict['name'] = name
                        if surname:
                            row_dict['surname'] = surname
                        if patronymic:
                            row_dict['patronymic'] = patronymic
                        # If parsing failed but we have a value, use it as surname
                        if not surname and counterparty_str:
                            row_dict['surname'] = counterparty_str
            
            rows.append(row_dict)
        
        if not rows:
            raise ValueError("Excel file is empty or has no data rows.")
        
        return rows
    
    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {str(e)}")


def import_customers_from_data(data_rows, skip_duplicates=True, skip_empty=True, default_place='Unknown'):
    """Import customers from parsed data rows"""
    imported = 0
    skipped = 0
    created_new = 0
    updated_existing = 0
    errors = []
    
    for i, row in enumerate(data_rows, start=2):  # Start at 2 (row 1 is header)
        try:
            # Get name, surname, and patronymic (required)
            name = row.get('name', '').strip() if row.get('name') else ''
            surname = row.get('surname', '').strip() if row.get('surname') else ''
            patronymic = row.get('patronymic', '').strip() if row.get('patronymic') else None
            if not patronymic:
                patronymic = None
            
            # If we have counterparty but no name/surname, try to parse it
            if (not name or not surname) and 'counterparty' in row:
                counterparty_value = row.get('counterparty', '')
                if counterparty_value:
                    # Convert to string and clean
                    counterparty_str = str(counterparty_value).strip()
                    if counterparty_str and counterparty_str.lower() not in ['none', 'null', 'nan', '']:
                        parsed_name, parsed_surname, parsed_patronymic = parse_counterparty_name(counterparty_str)
                        if parsed_name:
                            name = parsed_name
                        if parsed_surname:
                            surname = parsed_surname
                        if parsed_patronymic:
                            patronymic = parsed_patronymic
                        # If still no surname but we have counterparty, use it as surname
                        if not surname and counterparty_str:
                            surname = counterparty_str
            
            # Get place (use default if not provided)
            place = row.get('place', '').strip() or default_place
            phone = row.get('phone', '').strip() or None
            address = row.get('address', '').strip() or None
            
            # Validate required fields
            if not surname:
                # If skip_empty is True and counterparty is empty, just skip this row
                if skip_empty:
                    counterparty_val = row.get('counterparty', '')
                    if not counterparty_val or str(counterparty_val).strip() == '':
                        skipped += 1
                        continue
                
                # Try to get counterparty value for error message
                counterparty_info = ""
                if 'counterparty' in row:
                    counterparty_val = row.get('counterparty', '')
                    if counterparty_val:
                        counterparty_info = f" (Контрагент: '{str(counterparty_val)[:50]}')"
                    else:
                        counterparty_info = " (Контрагент is empty)"
                else:
                    # Check if there are any other columns that might have the name
                    all_keys = list(row.keys())
                    counterparty_info = f" (Available columns: {', '.join(all_keys[:5])})"
                errors.append(f"Row {i}: Surname is required{counterparty_info}")
                continue
            
            # If no name provided, use empty string
            if not name:
                name = ""
            
            # Create customer
            if skip_duplicates:
                # Check for duplicates first - if exists, skip
                query = Customer.objects.filter(
                    name=name,
                    surname=surname,
                    place=place
                )
                # Handle patronymic (can be None or empty)
                if patronymic:
                    query = query.filter(patronymic=patronymic)
                else:
                    query = query.filter(patronymic__isnull=True) | query.filter(patronymic='')
                
                if query.exists():
                    skipped += 1
                    continue
                
                # No duplicate found, create new customer
                customer = Customer.objects.create(
                    name=name,
                    surname=surname,
                    patronymic=patronymic,
                    place=place,
                    phone=phone,
                    address=address
                )
                imported += 1
                created_new += 1
            else:
                # skip_duplicates is False - try to create, but handle unique constraint
                try:
                    customer = Customer.objects.create(
                        name=name,
                        surname=surname,
                        patronymic=patronymic,
                        place=place,
                        phone=phone,
                        address=address
                    )
                    imported += 1
                    created_new += 1
                except IntegrityError:
                    # If unique constraint violation, customer already exists
                    # Get the existing customer and update if needed
                    try:
                        # Build query to handle patronymic being None or empty string
                        query = Customer.objects.filter(
                            name=name,
                            surname=surname,
                            place=place
                        )
                        # Handle patronymic (can be None, empty string, or a value)
                        if patronymic:
                            query = query.filter(patronymic=patronymic)
                        else:
                            # Match None or empty string
                            query = query.filter(patronymic__isnull=True) | query.filter(patronymic='')
                        
                        customer = query.first()
                        if not customer:
                            # Shouldn't happen if IntegrityError was raised, but handle it
                            errors.append(f"Row {i}: Unique constraint violation but customer not found")
                            continue
                        
                        # Update phone/address if provided
                        updated = False
                        if phone and not customer.phone:
                            customer.phone = phone
                            updated = True
                        if address and not customer.address:
                            customer.address = address
                            updated = True
                        if updated:
                            customer.save()
                            updated_existing += 1
                        # Count as imported since user wants to import all rows
                        imported += 1
                    except Exception as e2:
                        errors.append(f"Row {i}: Error updating existing customer: {str(e2)}")
                        continue
                except Exception as e:
                    # Other exceptions (not IntegrityError)
                    errors.append(f"Row {i}: Could not create customer: {str(e)}")
                    continue
        
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")
    
    return {
        'imported': imported,
        'skipped': skipped,
        'created_new': created_new,
        'updated_existing': updated_existing,
        'errors': errors
    }

